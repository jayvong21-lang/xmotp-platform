#!/usr/bin/env python3
"""xmotp.com - 厦门高端办公空间房源管理平台"""
import os, sys, csv, io, sqlite3, hashlib
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, jsonify, g, flash
from sync_public_properties import sync_public_properties

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'xmotp-secret-' + os.urandom(12).hex())
DATABASE = os.environ.get('XMOTP_DATABASE', os.path.join(os.path.dirname(__file__), 'xmotp.db'))

COLUMN_MAP = {
    '区域': 'district', 'district': 'district',
    '楼盘': 'building_name', '楼盘名': 'building_name', 'building_name': 'building_name',
    '面积': 'area_m2', '面积㎡': 'area_m2', '面积(㎡)': 'area_m2', 'area_m2': 'area_m2',
    '日租金': 'rent_per_day', '日租金(元/㎡)': 'rent_per_day', 'rent_per_day': 'rent_per_day',
    '月租金': 'total_rent_month', '月租': 'total_rent_month', 'total_rent_month': 'total_rent_month',
    '物业费': 'property_fee', 'property_fee': 'property_fee',
    '楼层': 'floor_info', 'floor_info': 'floor_info',
    '装修': 'decoration', 'decoration': 'decoration',
    '停车位': 'parking', '停车': 'parking', 'parking': 'parking',
    '合同到期': 'lease_expiry', 'lease_expiry': 'lease_expiry',
    '来源': 'source', '渠道': 'source', 'source': 'source',
    '联系人': 'contact_name', 'contact_name': 'contact_name',
    '联系电话': 'contact_phone', '电话': 'contact_phone', 'contact_phone': 'contact_phone',
    '状态': 'status', 'status': 'status',
    '备注': 'notes', 'notes': 'notes',
}

# ── 数据库 ────────────────────────────────────────────

def get_db():
    if 'db' not in g:
        g.db = sqlite3.connect(DATABASE)
        g.db.row_factory = sqlite3.Row
        g.db.execute("PRAGMA journal_mode=WAL")
    return g.db

@app.teardown_appcontext
def close_db(e=None):
    g.pop('db', None)

def init_db():
    db = sqlite3.connect(DATABASE)
    db.executescript("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            display_name TEXT DEFAULT '',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS listings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            district TEXT NOT NULL DEFAULT '',
            building_name TEXT NOT NULL DEFAULT '',
            area_m2 REAL DEFAULT 0,
            rent_per_day REAL DEFAULT 0,
            total_rent_month REAL DEFAULT 0,
            floor_info TEXT DEFAULT '',
            decoration TEXT DEFAULT '精装',
            property_fee REAL DEFAULT 0,
            parking TEXT DEFAULT '',
            lease_expiry TEXT DEFAULT '',
            source TEXT DEFAULT '',
            contact_name TEXT DEFAULT '',
            contact_phone TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            status TEXT DEFAULT '在租',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS listings_duplicates_archive (
            archive_id INTEGER PRIMARY KEY AUTOINCREMENT,
            original_id INTEGER UNIQUE NOT NULL,
            district TEXT, building_name TEXT, area_m2 REAL, rent_per_day REAL,
            total_rent_month REAL, floor_info TEXT, decoration TEXT,
            property_fee REAL, parking TEXT, lease_expiry TEXT, source TEXT,
            contact_name TEXT, contact_phone TEXT, notes TEXT, status TEXT,
            created_at DATETIME, updated_at DATETIME,
            archived_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            archive_reason TEXT DEFAULT 'duplicate district/building/area'
        );
    """)
    # 历史数据迁移：同区域、楼盘、面积只保留更新时间最新（同时间保留 id 最大）的一条。
    db.execute("""
        INSERT OR IGNORE INTO listings_duplicates_archive (
            original_id, district, building_name, area_m2, rent_per_day,
            total_rent_month, floor_info, decoration, property_fee, parking,
            lease_expiry, source, contact_name, contact_phone, notes, status,
            created_at, updated_at
        )
        SELECT id, district, building_name, area_m2, rent_per_day,
            total_rent_month, floor_info, decoration, property_fee, parking,
            lease_expiry, source, contact_name, contact_phone, notes, status,
            created_at, updated_at
        FROM listings
        WHERE EXISTS (
            SELECT 1 FROM listings newer
            WHERE newer.district = listings.district
              AND newer.building_name = listings.building_name
              AND newer.area_m2 = listings.area_m2
              AND (
                  datetime(newer.updated_at) > datetime(listings.updated_at)
                  OR (datetime(newer.updated_at) = datetime(listings.updated_at) AND newer.id > listings.id)
              )
        )
    """)
    db.execute("""
        DELETE FROM listings
        WHERE EXISTS (
            SELECT 1 FROM listings newer
            WHERE newer.district = listings.district
              AND newer.building_name = listings.building_name
              AND newer.area_m2 = listings.area_m2
              AND (
                  datetime(newer.updated_at) > datetime(listings.updated_at)
                  OR (datetime(newer.updated_at) = datetime(listings.updated_at) AND newer.id > listings.id)
              )
        )
    """)
    db.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_listings_dedup
        ON listings(district, building_name, area_m2)
    """)
    # 默认管理员账号: admin / admin123
    pwd = hashlib.sha256('admin123'.encode()).hexdigest()
    db.execute("INSERT OR IGNORE INTO users (username, password_hash, display_name) VALUES (?, ?, ?)",
               ('admin', pwd, '管理员'))
    db.commit()
    db.close()
    sync_properties_if_configured()

def sync_properties_if_configured():
    default_output = '/var/www/longc-es/js/properties-data.js'
    if not os.path.isdir(os.path.dirname(default_output)):
        default_output = ''
    output = os.environ.get('LONGC_PROPERTIES_OUTPUT', default_output).strip()
    if not output:
        return 0
    try:
        return sync_public_properties(DATABASE, output)
    except Exception:
        app.logger.exception('同步龙溪官网房源失败')
        return -1

# ── 登录 ──────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if 'user' not in session:
            return redirect(url_for('login'))
        return f(*a, **kw)
    return decorated

def api_login_required(f):
    @wraps(f)
    def decorated(*a, **kw):
        if 'user' not in session:
            return jsonify({'error': 'authentication required'}), 401
        return f(*a, **kw)
    return decorated

def normalize_row(row):
    """把中英文表头、空值和常见 Excel 数值统一为数据库字段。"""
    clean = {}
    for key, value in row.items():
        if key is None:
            continue
        key = str(key).strip()
        db_col = COLUMN_MAP.get(key, key.lower().replace(' ', '_'))
        if value is None or str(value).strip().lower() == 'nan':
            clean[db_col] = ''
        else:
            clean[db_col] = str(value).strip()
    return clean

def number(value):
    try:
        if value is None or str(value).strip() == '':
            return 0
        return float(str(value).replace(',', '').strip())
    except (TypeError, ValueError):
        return 0

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        db = get_db()
        u = db.execute("SELECT * FROM users WHERE username=?", (request.form['username'],)).fetchone()
        if u and u['password_hash'] == hashlib.sha256(request.form['password'].encode()).hexdigest():
            session['user'] = {'id': u['id'], 'name': u['display_name'] or u['username']}
            return redirect(url_for('index'))
        return render_template('login.html', error='用户名或密码错误')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ── 房源管理 ──────────────────────────────────────────

@app.route('/')
@login_required
def index():
    db = get_db()
    # 筛选参数
    district = request.args.get('district', '').strip()
    status = request.args.get('status', '在租').strip()
    keyword = request.args.get('keyword', '').strip()
    area_min = request.args.get('area_min', '').strip()
    area_max = request.args.get('area_max', '').strip()
    rent_min = request.args.get('rent_min', '').strip()
    rent_max = request.args.get('rent_max', '').strip()
    page = max(1, int(request.args.get('page', '1') or '1'))
    sort = request.args.get('sort', 'updated').strip()
    days = request.args.get('days', '').strip()

    conditions = []
    params = []

    if status != '全部':
        conditions.append("status = ?")
        params.append(status)

    if district:
        conditions.append("district = ?")
        params.append(district)

    if days and days.isdigit() and int(days) > 0:
        conditions.append("created_at >= datetime('now', ?)")
        params.append(f'-{days} days')

    if keyword:
        conditions.append("(building_name LIKE ? OR contact_name LIKE ? OR contact_phone LIKE ? OR notes LIKE ?)")
        kw = f'%{keyword}%'
        params.extend([kw, kw, kw, kw])

    if area_min:
        conditions.append("area_m2 >= ?")
        params.append(float(area_min))
    if area_max:
        conditions.append("area_m2 <= ?")
        params.append(float(area_max))

    if rent_min:
        conditions.append("rent_per_day >= ?")
        params.append(float(rent_min))
    if rent_max:
        conditions.append("rent_per_day <= ?")
        params.append(float(rent_max))

    where = f"WHERE {' AND '.join(conditions)}" if conditions else ""
    per_page = 20
    offset = (page - 1) * per_page

    order_map = {'updated': 'updated_at DESC', 'area': 'area_m2 DESC', 'rent': 'rent_per_day ASC', 'created': 'created_at DESC'}
    order = order_map.get(sort, 'updated_at DESC')

    total = db.execute(f"SELECT COUNT(*) FROM listings {where}", params).fetchone()[0]
    listings = db.execute(
        f"SELECT * FROM listings {where} ORDER BY {order} LIMIT ? OFFSET ?",
        params + [per_page, offset]
    ).fetchall()

    total_pages = max(1, (total + per_page - 1) // per_page)

    # 获取所有区域和状态用于筛选
    districts = [r['district'] for r in db.execute("SELECT DISTINCT district FROM listings WHERE district != '' ORDER BY district").fetchall()]

    stats = dict(db.execute(
        "SELECT status, COUNT(*) FROM listings GROUP BY status"
    ).fetchall())

    def page_url(p):
        args = request.args.copy()
        args['page'] = p
        from urllib.parse import urlencode
        return '?' + urlencode(args.to_dict(flat=False), doseq=True)

    return render_template('index.html',
        listings=listings, districts=districts,
        stats=stats, page=page, total_pages=total_pages, total=total,
        cur_district=district, cur_status=status, cur_keyword=keyword,
        cur_area_min=area_min, cur_area_max=area_max,
        cur_rent_min=rent_min, cur_rent_max=rent_max,
        cur_sort=sort, cur_days=days,
        page_url=page_url,
    )

@app.route('/listing/new', methods=['GET', 'POST'])
@login_required
def listing_new():
    if request.method == 'POST':
        db = get_db()
        duplicate = db.execute(
            "SELECT id FROM listings WHERE district=? AND building_name=? AND area_m2=?",
            (request.form.get('district', ''), request.form.get('building_name', ''),
             number(request.form.get('area_m2', 0)))
        ).fetchone()
        if duplicate:
            return render_template(
                'form.html', listing=request.form,
                error='该区域、楼盘和面积的房源已存在，请编辑原记录。'
            ), 409
        db.execute("""
            INSERT INTO listings (district, building_name, area_m2, rent_per_day, total_rent_month,
                floor_info, decoration, property_fee, parking, lease_expiry,
                source, contact_name, contact_phone, notes, status)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            request.form.get('district', ''),
            request.form.get('building_name', ''),
            float(request.form.get('area_m2', 0) or 0),
            float(request.form.get('rent_per_day', 0) or 0),
            float(request.form.get('total_rent_month', 0) or 0),
            request.form.get('floor_info', ''),
            request.form.get('decoration', '精装'),
            float(request.form.get('property_fee', 0) or 0),
            request.form.get('parking', ''),
            request.form.get('lease_expiry', ''),
            request.form.get('source', ''),
            request.form.get('contact_name', ''),
            request.form.get('contact_phone', ''),
            request.form.get('notes', ''),
            '在租'
        ))
        db.commit()
        sync_properties_if_configured()
        return redirect(url_for('index'))
    return render_template('form.html', listing=None)

@app.route('/listing/<int:id>/edit', methods=['GET', 'POST'])
@login_required
def listing_edit(id):
    db = get_db()
    listing = db.execute("SELECT * FROM listings WHERE id=?", (id,)).fetchone()
    if not listing:
        return "房源不存在", 404
    if request.method == 'POST':
        duplicate = db.execute(
            "SELECT id FROM listings WHERE district=? AND building_name=? AND area_m2=? AND id!=?",
            (request.form.get('district', ''), request.form.get('building_name', ''),
             number(request.form.get('area_m2', 0)), id)
        ).fetchone()
        if duplicate:
            return render_template(
                'form.html', listing=listing,
                error='修改后会与现有房源重复，请调整区域、楼盘或面积。'
            ), 409
        db.execute("""
            UPDATE listings SET district=?, building_name=?, area_m2=?, rent_per_day=?,
                total_rent_month=?, floor_info=?, decoration=?, property_fee=?, parking=?,
                lease_expiry=?, source=?, contact_name=?, contact_phone=?, notes=?,
                status=?, updated_at=CURRENT_TIMESTAMP
            WHERE id=?
        """, (
            request.form.get('district', ''),
            request.form.get('building_name', ''),
            float(request.form.get('area_m2', 0) or 0),
            float(request.form.get('rent_per_day', 0) or 0),
            float(request.form.get('total_rent_month', 0) or 0),
            request.form.get('floor_info', ''),
            request.form.get('decoration', '精装'),
            float(request.form.get('property_fee', 0) or 0),
            request.form.get('parking', ''),
            request.form.get('lease_expiry', ''),
            request.form.get('source', ''),
            request.form.get('contact_name', ''),
            request.form.get('contact_phone', ''),
            request.form.get('notes', ''),
            request.form.get('status', '在租'),
            id
        ))
        db.commit()
        sync_properties_if_configured()
        return redirect(url_for('index'))
    return render_template('form.html', listing=listing)

@app.route('/listing/<int:id>/delete', methods=['POST'])
@login_required
def listing_delete(id):
    db = get_db()
    db.execute("UPDATE listings SET status='已下架', updated_at=CURRENT_TIMESTAMP WHERE id=?", (id,))
    db.commit()
    sync_properties_if_configured()
    return redirect(url_for('index'))

@app.route('/listing/<int:id>/detail')
@login_required
def listing_detail(id):
    db = get_db()
    listing = db.execute("SELECT * FROM listings WHERE id=?", (id,)).fetchone()
    if not listing:
        return "房源不存在", 404
    return render_template('detail.html', listing=listing)

@app.route('/export')
@login_required
def export():
    db = get_db()
    rows = db.execute("SELECT * FROM listings WHERE status='在租' ORDER BY district, building_name").fetchall()
    import csv, io
    si = io.StringIO()
    w = csv.writer(si)
    w.writerow(['区域', '楼盘名', '面积(㎡)', '日租金(元/㎡)', '月租金', '楼层', '装修', '物业费', '停车位', '合同到期', '来源', '联系人', '电话', '备注'])
    for r in rows:
        w.writerow([r['district'], r['building_name'], r['area_m2'], r['rent_per_day'],
                    r['total_rent_month'], r['floor_info'], r['decoration'], r['property_fee'],
                    r['parking'], r['lease_expiry'], r['source'], r['contact_name'],
                    r['contact_phone'], r['notes']])
    output = si.getvalue().encode('utf-8-sig')
    return app.response_class(output, mimetype='text/csv',
                              headers={'Content-Disposition': 'attachment; filename=房源导出.csv'})

# ── 批量导入 ──────────────────────────────────────────

CSV_COLUMNS = '区域,楼盘名,面积(㎡),日租金(元/㎡),月租金,楼层,装修,物业费,停车位,合同到期,来源,联系人,电话,备注'

@app.route('/import', methods=['GET', 'POST'])
@login_required
def import_listings():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            return render_template('import.html', error='请选择文件')

        filename = file.filename.lower()
        if not (filename.endswith('.csv') or filename.endswith('.xlsx')):
            return render_template('import.html', error='仅支持 CSV 或 Excel(.xlsx) 文件')

        try:
            content = file.read()
            if filename.endswith('.csv'):
                # 尝试 UTF-8 和 GBK
                try:
                    text = content.decode('utf-8-sig')
                except UnicodeDecodeError:
                    text = content.decode('gbk')
                reader = csv.DictReader(io.StringIO(text))
                rows = list(reader)
            else:
                # Excel: 需要 openpyxl
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    return render_template('import.html', error='服务器未安装 Excel 支持，请使用 CSV 格式')
                wb = load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
                wb.close()

            if not rows:
                return render_template('import.html', error='文件中没有数据')

        except Exception as e:
            return render_template('import.html', error=f'文件解析失败: {str(e)}')

        db = get_db()
        inserted, updated, skipped = 0, 0, 0

        for row in rows:
            clean = normalize_row(row)

            building = clean.get('building_name', '')
            district = clean.get('district', '')
            if not building:
                skipped += 1
                continue

            area = number(clean.get('area_m2', 0))
            rent = number(clean.get('rent_per_day', 0))
            total_rent = number(clean.get('total_rent_month', 0))

            # 去重: 同区域+同楼盘+同面积 → 更新
            exist = db.execute(
                "SELECT id FROM listings WHERE district=? AND building_name=? AND area_m2=?",
                (district, building, area)
            ).fetchone()

            if exist:
                db.execute("""UPDATE listings SET
                    rent_per_day=?, total_rent_month=?, floor_info=?, decoration=?,
                    property_fee=?, parking=?, lease_expiry=?, source=?,
                    contact_name=?, contact_phone=?, notes=?,
                    status='在租', updated_at=CURRENT_TIMESTAMP
                    WHERE id=?""", (
                    rent, total_rent,
                    clean.get('floor_info', ''), clean.get('decoration', '精装'),
                    number(clean.get('property_fee', 0)), clean.get('parking', ''),
                    clean.get('lease_expiry', ''), clean.get('source', ''),
                    clean.get('contact_name', ''), clean.get('contact_phone', ''),
                    clean.get('notes', ''), exist['id']
                ))
                updated += 1
            else:
                db.execute("""INSERT INTO listings
                    (district, building_name, area_m2, rent_per_day, total_rent_month,
                     floor_info, decoration, property_fee, parking, lease_expiry,
                     source, contact_name, contact_phone, notes, status)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""", (
                    district, building, area, rent, total_rent,
                    clean.get('floor_info', ''), clean.get('decoration', '精装'),
                    number(clean.get('property_fee', 0)), clean.get('parking', ''),
                    clean.get('lease_expiry', ''), clean.get('source', ''),
                    clean.get('contact_name', ''), clean.get('contact_phone', ''),
                    clean.get('notes', ''), '在租'
                ))
                inserted += 1

        db.commit()
        sync_properties_if_configured()

        result = f'导入完成：新增 {inserted} 条，更新 {updated} 条'
        if skipped: result += f'，跳过 {skipped} 条（无楼盘名）'
        return render_template('import.html', success=result)

    return render_template('import.html')

@app.route('/import/template')
@login_required
def import_template():
    output = io.StringIO()
    output.write(CSV_COLUMNS + '\n')
    output.write('思明区,世茂海峡大厦,500,3.5,52500,15层/共32层 朝南,精装,15,地下停车 800元/月,2026-12-31,58同城,张先生,13800001111,大堂气派\n')
    output.write('湖里区,国际航运中心,300,4.0,36000,20层,精装,18,免费停车,2027-06-30,安居客,李女士,13900002222,\n')
    return app.response_class(
        output.getvalue().encode('utf-8-sig'), mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=房源导入模板.csv'}
    )

# ── 查询 API（包含联系人信息，必须登录） ───────────────

@app.route('/api/listings')
@api_login_required
def api_listings():
    db = get_db()
    limit = min(max(request.args.get('limit', 200, type=int), 1), 500)
    offset = max(request.args.get('offset', 0, type=int), 0)
    status = request.args.get('status', '').strip()
    district = request.args.get('district', '').strip()
    conditions, params = [], []
    if status:
        conditions.append('status=?')
        params.append(status)
    if district:
        conditions.append('district=?')
        params.append(district)
    where = ' WHERE ' + ' AND '.join(conditions) if conditions else ''
    rows = db.execute(
        'SELECT * FROM listings' + where + ' ORDER BY updated_at DESC LIMIT ? OFFSET ?',
        params + [limit, offset]
    ).fetchall()
    return jsonify([dict(row) for row in rows])

@app.route('/api/stats')
@api_login_required
def api_stats():
    db = get_db()
    stats = {
        'total': db.execute('SELECT COUNT(*) FROM listings').fetchone()[0],
        'active': db.execute("SELECT COUNT(*) FROM listings WHERE status='在租'").fetchone()[0],
        'today_new': db.execute("SELECT COUNT(*) FROM listings WHERE date(created_at)=date('now','localtime')").fetchone()[0],
        'today_updated': db.execute("SELECT COUNT(*) FROM listings WHERE date(updated_at)=date('now','localtime') AND date(created_at)!=date('now','localtime')").fetchone()[0],
    }
    return jsonify(stats)

# ── 启动 ──────────────────────────────────────────────

init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5100, debug=False)
